
import openpyxl
import psycopg2


SCHEMA_NAME = 'public'
TABLE_NAME = 'mst_stock'
STOCK_EXCHANGE_CODE = 'IDX'
FILEPATH = './a.xlsx'

db_conn = psycopg2.connect(
    host='satao.db.elephantsql.com',
    port=5432,
    dbname='xgpbnhoy',
    user='xgpbnhoy',
    password='DTAm6UYROwh1iDRZstKN1hXyJlAgC7-s'
)

db_cursor = db_conn.cursor()

selector_query = f'''
    SELECT column_name
    FROM information_schema.columns
    WHERE table_schema = '{SCHEMA_NAME}'
        AND table_name = '{TABLE_NAME}';
'''

db_cursor.execute(selector_query)

headings = tuple(map(lambda x: x[0], db_cursor.fetchall()))

selector_query = f'''
    SELECT *
    FROM {TABLE_NAME}
    WHERE stock_exchange_code = '{STOCK_EXCHANGE_CODE}'
    ORDER BY stock_code;
'''

db_cursor.execute(selector_query)

wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()

sheet.row_dimensions[1].font = openpyxl.styles.Font(bold = True)

# Spreadsheet row and column indexes start at 1
# so we use "start = 1" in enumerate so
# we don't need to add 1 to the indexes.
for colno, heading in enumerate(headings, start = 1):
    sheet.cell(row = 1, column = colno).value = heading

# This time we use "start = 2" to skip the heading row.
for rowno, row in enumerate(db_cursor.fetchall(), start = 2):
    for colno, cell_value in enumerate(row, start = 1):
        sheet.cell(row = rowno, column = colno).value = cell_value

wb.save(FILEPATH)

db_cursor.close()